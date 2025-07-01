import openai
import pandas as pd
from azure.ai.translation.text import TextTranslationClient
from azure.core.credentials import AzureKeyCredential
from openai import AzureOpenAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# === Step 1: Insert Your API Keys ===
AZURE_KEY = "YOUR_AZURE_KEY_HERE"
AZURE_REGION = "YOUR_AZURE_REGION"  # e.g., "eastus"
OPENAI_KEY = "YOUR_OPENAI_KEY_HERE"

# Set OpenAI API key
openai.api_key = OPENAI_KEY

# === Step 2: Dialect Code to Language Mapping ===
dialect_lang_map = {
    "ARG": "es-ar",
    "CHL": "es-cl",
    "COL": "es-co",
    "MEX": "es-mx",
    "BRA": "pt-br"
}

# === Step 3: Load the labeled input data ===
input_file = "labeled_output.txt"
rows = []

with open(input_file, "r", encoding="utf-8") as f:
    for line in f:
        if "=>" in line:
            sentence, code = line.strip().split("=>")
            sentence = sentence.strip()
            code = code.strip().upper()
            rows.append((sentence, code))

# === Step 4: Updated Azure Translate Function with Error Catching ===
def azure_translate(text, source_lang):
    endpoint = "YOUR_AZURE_ENDPOINT_HERE"
    region = "YOUR_AZURE_REGION"
    api_key = AZURE_KEY
    
    # Create the translation client
    credential = AzureKeyCredential(api_key)
    translator = TextTranslationClient(endpoint=endpoint, region=region, credential=credential)

    try:
        # Perform the translation
        result = translator.translate(content=text, to=["en"], from_parameter=source_lang)
        
        # Extract the translation result
        translation = result[0].translations[0].text
        return translation
    except Exception as e:
        # Log and print the error for debugging purposes
        print(f"Error in Azure translation for '{text}': {e}")
        return "ERROR"

# === Step 5: Updated OpenAI Translate Function with Error Catching ===
def openai_translate(text, source_lang):
    azure_endpoint = "YOUR_OPENAI_ENDPOINT_HERE"
    subscription_key = "YOUR_OPENAI_KEY_HERE"
    api_version = "2024-12-01-preview"
    model_name = "gpt-4o"
    deployment = "gpt-4o"

    # Create the AzureOpenAI client
    client = AzureOpenAI(
        api_version=api_version,
        azure_endpoint=azure_endpoint,
        api_key=subscription_key
    )

    prompt = f"Translate the following sentence from {source_lang} to English:\n\n{text}"

    try:
        # Call the Azure OpenAI model for translation
        response = client.chat.completions.create(
            messages=[
                {"role": "system", "content": "You are a professional translator."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=100,
            temperature=0.3,
            model=deployment
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Error in OpenAI translation for '{text}': {e}")
        return "ERROR"

# === Step 6: Translation Pipeline with Error Handling ===
output_data = []
for sentence, code in rows:
    lang_code = dialect_lang_map.get(code, "es")
    
    try:
        # First, try Azure translation
        azure_result = azure_translate(sentence, lang_code)
    except Exception as e:
        azure_result = f"Azure Error: {e}"

    try:
        # Then, try OpenAI translation
        openai_result = openai_translate(sentence, lang_code)
    except Exception as e:
        openai_result = f"OpenAI Error: {e}"
    
    # Append the result to the output data
    output_data.append([sentence, code, azure_result, openai_result])

# === Step 7: Save to Excel ===
df = pd.DataFrame(output_data, columns=["Original", "Dialect", "Azure Translation", "OpenAI Translation"])
excel_filename = "translated_output.xlsx"
df.to_excel(excel_filename, index=False)

print("✅ Translations completed. Output saved to 'translated_output.xlsx'")

# === Step 8: Apply Red Color for Errors in Excel ===
# Load the workbook and select the active sheet
wb = load_workbook(excel_filename)
ws = wb.active

# Define the red background and white bold text style
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
white_bold_font = Font(color="FFFFFF", bold=False)

# Loop through the cells to find errors and apply styling
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):  # Skipping the header row
    for cell in row:
        if cell.value == "ERROR" or isinstance(cell.value, str) and "Error" in cell.value:
            cell.fill = red_fill
            cell.font = white_bold_font

# Save the modified Excel file with red error cells
wb.save(excel_filename)
print(f"✅ Excel file updated with error highlights. Output saved to '{excel_filename}'")